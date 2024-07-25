import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def read_spreadsheets(file_path1, file_path2, output_path):
    df1 = pd.read_excel(file_path1)
    df2 = pd.read_excel(file_path2)
    
    #sort_values ordena os valores e reset_index reordena a numeração das linhas [0,1,2,...]
    df1_data = df1.sort_values(by=df1.columns.tolist()).reset_index(drop=True).fillna('')
    df2_data = df2.sort_values(by=df2.columns.tolist()).reset_index(drop=True).fillna('')

    # Create a writer object to write the results
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df1.to_excel(writer, index=False)

    # Load the workbook and sheet
    wb = load_workbook(output_path)
    ws = wb.active

    # Define colors using ARGB format
    IndianRed_fill = PatternFill(start_color="FFCD5C5C", end_color="FFCD5C5C", fill_type="solid") #Header Columns with error
    LightCoral_fill = PatternFill(start_color="FFF08080", end_color="FFF08080", fill_type="solid") #Columns with error
    DodgerBlue_fill = PatternFill(start_color="FF1E90FF", end_color="FF1E90FF", fill_type="solid") #Empty columns
    Khaki_fill = PatternFill(start_color="FFF0E68C", end_color="FFF0E68C", fill_type="solid") #Is no header on spreadsheet to compare

    for column in df1_data:
        header_cell = ws.cell(row=1, column=df1.columns.get_loc(column) + 1)
        column_has_difference = False

        if column in df2_data:
            for index, line in df1_data[column].items():
                print(index, line)
                if df1_data[column][index] != df2_data[column][index]:
                    print(f'{df1_data[column][index]} != {df2_data[column][index]}')
                    print()
                    cell = ws.cell(row=index+2, column=df1.columns.get_loc(column)+1)
                    cell.fill = LightCoral_fill
                    header_cell.fill = IndianRed_fill
                    
        else:
            column_has_difference = True
        
            if column_has_difference:
                header_cell.fill = Khaki_fill
    
    wb.save(output_path)
if __name__ == '__main__':
    file_path1 = './Excel1.xlsx'
    file_path2 = './Excel2.xlsx'
    output_path = './Result_excel.xlsx'
    read_spreadsheets(file_path1, file_path2, output_path)

